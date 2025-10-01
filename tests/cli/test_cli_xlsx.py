from __future__ import annotations

from pathlib import Path
from typer.testing import CliRunner
from json_to_excel_converter.cli import app
import openpyxl


def project_root() -> Path:
    return Path(__file__).resolve().parents[2]


def test_xlsx_basic(tmp_path: Path):
    runner = CliRunner()
    src = project_root() / "examples" / "ads_small.json"
    dst = tmp_path / "out.xlsx"

    result = runner.invoke(app, [str(src), str(dst), "--root", "items", "--sheet-name", "Items", "--first-column", "id"])
    assert result.exit_code == 0, result.output
    assert dst.exists()

    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    assert ws.title == "Items"

    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header[0] == "id"
    assert any(str(c).startswith("summary.") for c in header)


def test_xlsx_include_and_order(tmp_path: Path):
    runner = CliRunner()
    src = project_root() / "examples" / "ads_small.json"
    dst = tmp_path / "out.xlsx"

    result = runner.invoke(
        app,
        [
            str(src),
            str(dst),
            "--root",
            "items",
            "--sheet-name",
            "Items",
            "--first-column",
            "id",
            "--include",
            "summary",
            "--include",
            "details",
        ],
    )
    assert result.exit_code == 0, result.output
    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(max_row=1))]
    assert header[0] == "id"
    assert all(c == "id" or str(c).startswith("summary.") or str(c).startswith("details.") for c in header)
    assert str(header[1]).startswith("summary.")
